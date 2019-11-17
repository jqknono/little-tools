from datetime import date, datetime, timedelta
from xlutils.copy import copy
import openpyxl
from functions import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color 
from copy import copy
from math import log10, floor

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
            'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
            

def assign_value(ws, row, column, value):
    ws[f'{column}{row}'] = value
    ws[f'{column}{row}'].font = copy(ws[f'{column}{row-50}'].font)
    ws[f'{column}{row}'].alignment = copy(ws[f'{column}{row-50}'].alignment)
    ws[f'{column}{row}'].fill  = copy(ws[f'{column}{row-50}'].fill )
    ws[f'{column}{row}'].number_format = copy(ws[f'{column}{row-50}'].number_format)


yesterday = (today - timedelta(days=1)).strftime('%Y/%m/%d')


# template
# def save_shouqiang_data():
#     try:
#       get
#     except:
#       print
#       return
    
#     try:
#         handle
#     except:
#         print, return

#     try:
#         save
#     except:

    

def save_to_worksheet(ws):
    rows = int(ws.max_row)
    ###############
    # 查数据
    data = query_active_users()
    data1 = query_morrow_retentions()
    data2 = query_threedays_retentions()
    data3 = query_sevendays_retentions()
    data4 = query_launches_times()
    data5 = query_video()
    data6 = query_video_supplement()
    data7 = query_iap_users()
    data8 = query_unlock_business()
    data9 = query_open_map2_users()
    data10 = query_open_map3_users()
    data11 = query_open_event()
    data12 = query_claim_times()
    data13 = query_iap_money()
    data14 = calculate_iap_nums()
    data15 = calculate_iap_money_sum()
    data16 = query_ad_play_num_all()
    data17 = query_xi_back_gun()
    data18 = query_xi_ecpm()
    data19 = query_xi_ecpm(country_us)


    # 数据计算
    value_E = round_sig(data['昨天'][0] / data['昨天'][1])    #
    value_F = round_sig((3.3*(data5['昨天'][0] / data['昨天'][0])* (data18['昨天']/1000))+ (3.3*0.7*(data15/data['昨天'][0])))
    value_G_1 = round_sig(data18['昨天'])    # G列 昨日ecpm
    value_G_2 = round_sig(data18['前天'])    # G列 前日ecpm
    value_H_1 = round_sig(data19['昨天'])    # F列 昨日US-ecpm
    value_H_2 = round_sig(data19['前天'])    # F列 前日US-ecpm
    
    

    value_L = round_sig(data4['昨天'] / data['昨天'][0])     # L列日活启动次数
    value_M = round_sig(data5['昨天'][0] / data['昨天'][0])    # M列日活视频次数
    value_N = round_sig(data5['昨天'][0] / data5['昨天'][1])   # N列独立用户视频次数
    value_O = round_sig(data5['昨天'][1]/data['昨天'][0])     # O列视频用户覆盖率
    value_P = round_sig(data16['Double_Cash'] / data5['昨天'][1])       # P列doubleCash广告位独立用户人均视频数
    # Q列Offline广告位独立用户人均视频数
    value_Q = round_sig(data16['Offline']/data5['昨天'][1])
    value_R = round_sig(data16['AdClaim']/data5['昨天'][1])           # R列claim广告位独立用户人均视频数
    value_S1 = data16['FreeBonus_money']/data5['昨天'][1]
    value_S2 = data16['FreeBonus_gold']/data5['昨天'][1]
    value_S = round_sig(value_S1 + value_S2)                          # S列FreeBonus广告位独立用户人均视频数
    value_T = round_sig(data16['speedUp']/data5['昨天'][1])            # T列speed广告位独立用户人均视频数
    value_U = round_sig(data16['Slot']/data5['昨天'][1])               # U列spin广告位独立用户人均视频数
    value_V = round_sig(data16['skip_30_minutes']/data5['昨天'][1])    # V列skimtime广告位独立用户人均视频数
    value_W = round_sig(data16['cost_50%_off'] /data5['昨天'][1])       # W列50%discount广告位独立用户人均视频数
    value_X = round_sig(data6['昨天']/data['昨天'][0])                  # x列人均插频次数
    value_Y = round_sig((data5['昨天'][0]+data6['昨天'])/data['昨天'][0] )       # Y列人均视频+插频次数

    value_z = data14                                        # Z列内购次数

    value_AA = data7['昨天']                                 # AA列内购人数
    value_AB = round_sig(data15)                                       # AB列内购金额
    value_AD = round_sig(data14/data['昨天'][0])                        # AD列日活人均付费次数
    value_AE = round_sig(data7['昨天']/data['昨天'][0])                  # AE列日活付费率
    value_AF = round_sig(data15/data14)                                 # AF列单次付费价值
    value_AG = round_sig(data15/data['昨天'][0])                         # AG列日活人均付费金额

    value_AH = data8['昨天'][0]/data['昨天'][1]                # AH列新增用户解锁第5产业占比
    value_AI = data8['昨天'][1]/data['昨天'][1]                # AI列新增用户解锁第10产业占比
    value_AJ = round_sig(data12['昨天']/data['昨天'][0])                  # AJ列成功claim日活用户占比
    value_AK = data9['昨天']/data['昨天'][0]                   # AK列进入map2日活用户占比
    value_AL = data10['昨天']/data['昨天'][0]                  # AL列进入map3日活用户占比
    value_AM = data11['昨天']/data['昨天'][0]                  # AM列进入event日活用户占比
    value_AN = round_sig(data17['num_video_played']/data17['people_num_watch_video'])      # AN列M美国fb用户24小时视频数

    
    
    # 切换列数
    new_row = rows + 1
    row_G = new_row-1
    row_F = new_row-1    
    row_I = new_row-1
    row_J = new_row-3
    row_K = new_row-7
    row_AN= new_row-2

    # 填数据到excel
    yesterday_to_write = yesterday
    if yesterday[5] == '0':
        yesterday_to_write = yesterday[0:5]+yesterday[6:10]

    assign_value(ws, new_row, 'A', yesterday_to_write)
    assign_value(ws, new_row, 'B', 'iOS')
    assign_value(ws, new_row, 'C', data['昨天'][1])
    assign_value(ws, new_row, 'D', data['昨天'][0])
    assign_value(ws, new_row, 'E', value_E)
    assign_value(ws, new_row, 'F', value_F)
    assign_value(ws, new_row, 'G', value_G_1)
    assign_value(ws, row_G, 'G', value_G_2)
    assign_value(ws, new_row, 'H', value_H_1)
    assign_value(ws, row_F, 'H', value_H_2)
    assign_value(ws, row_I, 'I', (data1['昨天']/100))  # 转换为百分数
    assign_value(ws, row_J, 'J', (data2['昨天']/100))
    assign_value(ws, row_K, 'K', (data3['昨天'])/100)
    assign_value(ws, new_row, 'L', value_L)
    assign_value(ws, new_row, 'M', value_M)
    assign_value(ws, new_row, 'N', value_N)
    assign_value(ws, new_row, 'O',value_O)
    assign_value(ws, new_row, 'P', value_P)
    assign_value(ws, new_row, 'Q', value_Q)
    assign_value(ws, new_row, 'R', value_R)
    assign_value(ws, new_row, 'S', value_S)
    assign_value(ws, new_row, 'T', value_T)
    assign_value(ws, new_row, 'U', value_U)
    assign_value(ws, new_row, 'V', value_V)
    assign_value(ws, new_row, 'W', value_W)

    assign_value(ws, new_row, 'X', value_X)
    assign_value(ws, new_row, 'Y', value_Y)
    assign_value(ws, new_row, 'Z', value_z)
    assign_value(ws, new_row, 'AA', value_AA)
    assign_value(ws, new_row, 'AB', value_AB)
    assign_value(ws, new_row, 'AD', value_AD)
    assign_value(ws, new_row, 'AE', value_AE)
    assign_value(ws, new_row, 'AF', value_AF)
    assign_value(ws, new_row, 'AG', value_AG)
    assign_value(ws, new_row, 'AH', value_AH)
    assign_value(ws, new_row, 'AI', value_AI)
    assign_value(ws, new_row, 'AJ', value_AJ)
    assign_value(ws, new_row, 'AK', value_AK)
    assign_value(ws, new_row, 'AL', value_AL)
    assign_value(ws, new_row, 'AM', value_AM)
    assign_value(ws, row_AN, 'AN',value_AN)

def save_umeng_worksheet(ws):
    rows = int(ws.max_row)
    ###############
    # 查数据
    data = query_active_users()
    data1 = query_morrow_retentions()
    data2 = query_threedays_retentions()
    data3 = query_sevendays_retentions()
    data4 = query_launches_times()
    data5 = query_video()
    data6 = query_video_supplement()
    data7 = query_iap_users()
    data8 = query_unlock_business()
    data9 = query_open_map2_users()
    data10 = query_open_map3_users()
    data11 = query_open_event()
    data12 = query_claim_times()
    data13 = query_iap_money()
    data14 = calculate_iap_nums()
    data15 = calculate_iap_money_sum()
    data16 = query_ad_play_num_all()
    # data17 = query_xi_back_gun()
    data18 = query_xi_ecpm()
    # data19 = query_xi_ecpm(country_us)


    # 数据计算
    value_E = round_sig(data['昨天'][0] / data['昨天'][1])    #
    value_F = round_sig((3.3*(data5['昨天'][0] / data['昨天'][0])* (data18['昨天']/1000))+ (3.3*0.7*(data15/data['昨天'][0])))
    # value_G_1 = round_sig(data18['昨天'])    # G列 昨日ecpm
    # value_G_2 = round_sig(data18['前天'])    # G列 前日ecpm
    # value_H_1 = round_sig(data19['昨天'])    # F列 昨日US-ecpm
    # value_H_2 = round_sig(data19['前天'])    # F列 前日US-ecpm
    
    value_L = round_sig(data4['昨天'] / data['昨天'][0])     # L列日活启动次数
    value_M = round_sig(data5['昨天'][0] / data['昨天'][0])    # M列日活视频次数
    value_N = round_sig(data5['昨天'][0] / data5['昨天'][1])   # N列独立用户视频次数
    value_O = round_sig(data5['昨天'][1]/data['昨天'][0])     # O列视频用户覆盖率
    value_P = round_sig(data16['Double_Cash'] / data5['昨天'][1])       # P列doubleCash广告位独立用户人均视频数
    # Q列Offline广告位独立用户人均视频数
    value_Q = round_sig(data16['Offline']/data5['昨天'][1])
    value_R = round_sig(data16['AdClaim']/data5['昨天'][1])           # R列claim广告位独立用户人均视频数
    value_S1 = data16['FreeBonus_money']/data5['昨天'][1]
    value_S2 = data16['FreeBonus_gold']/data5['昨天'][1]
    value_S = round_sig(value_S1 + value_S2)                          # S列FreeBonus广告位独立用户人均视频数
    value_T = round_sig(data16['speedUp']/data5['昨天'][1])            # T列speed广告位独立用户人均视频数
    value_U = round_sig(data16['Slot']/data5['昨天'][1])               # U列spin广告位独立用户人均视频数
    value_V = round_sig(data16['skip_30_minutes']/data5['昨天'][1])    # V列skimtime广告位独立用户人均视频数
    value_W = round_sig(data16['cost_50%_off'] /data5['昨天'][1])       # W列50%discount广告位独立用户人均视频数
    value_X = round_sig(data6['昨天']/data['昨天'][0])                  # x列人均插频次数
    value_Y = round_sig((data5['昨天'][0]+data6['昨天'])/data['昨天'][0] )       # Y列人均视频+插频次数

    value_z = data14                                        # Z列内购次数

    value_AA = data7['昨天']                                 # AA列内购人数
    value_AB = round_sig(data15)                                       # AB列内购金额
    value_AD = round_sig(data14/data['昨天'][0])                        # AD列日活人均付费次数
    value_AE = round_sig(data7['昨天']/data['昨天'][0])                  # AE列日活付费率
    value_AF = round_sig(data15/data14)                                 # AF列单次付费价值
    value_AG = round_sig(data15/data['昨天'][0])                         # AG列日活人均付费金额

    value_AH = data8['昨天'][0]/data['昨天'][1]                # AH列新增用户解锁第5产业占比
    value_AI = data8['昨天'][1]/data['昨天'][1]                # AI列新增用户解锁第10产业占比
    value_AJ = round_sig(data12['昨天']/data['昨天'][0])                  # AJ列成功claim日活用户占比
    value_AK = data9['昨天']/data['昨天'][0]                   # AK列进入map2日活用户占比
    value_AL = data10['昨天']/data['昨天'][0]                  # AL列进入map3日活用户占比
    value_AM = data11['昨天']/data['昨天'][0]                  # AM列进入event日活用户占比
    # value_AN = round_sig(data17['num_video_played']/data17['people_num_watch_video'])      # AN列M美国fb用户24小时视频数

    
    
    # 切换列数
    new_row = rows + 1
    row_G = new_row-1
    row_F = new_row-1    
    row_I = new_row-1
    row_J = new_row-3
    row_K = new_row-7
    row_AN= new_row-2

    # 填数据到excel
    yesterday_to_write = yesterday
    if yesterday[5] == '0':
        yesterday_to_write = yesterday[0:5]+yesterday[6:10]

    assign_value(ws, new_row, 'A', yesterday_to_write)
    assign_value(ws, new_row, 'B', 'iOS')
    assign_value(ws, new_row, 'C', data['昨天'][1])
    assign_value(ws, new_row, 'D', data['昨天'][0])
    assign_value(ws, new_row, 'E', value_E)
    assign_value(ws, new_row, 'F', value_F)
    # assign_value(ws, new_row, 'G', value_G_1)
    # assign_value(ws, row_G, 'G', value_G_2)
    # assign_value(ws, new_row, 'H', value_H_1)
    # assign_value(ws, row_F, 'H', value_H_2)
    assign_value(ws, row_I, 'I', (data1['昨天']/100))  # 转换为百分数
    assign_value(ws, row_J, 'J', (data2['昨天']/100))
    assign_value(ws, row_K, 'K', (data3['昨天'])/100)
    assign_value(ws, new_row, 'L', value_L)
    assign_value(ws, new_row, 'M', value_M)
    assign_value(ws, new_row, 'N', value_N)
    assign_value(ws, new_row, 'O',value_O)
    assign_value(ws, new_row, 'P', value_P)
    assign_value(ws, new_row, 'Q', value_Q)
    assign_value(ws, new_row, 'R', value_R)
    assign_value(ws, new_row, 'S', value_S)
    assign_value(ws, new_row, 'T', value_T)
    assign_value(ws, new_row, 'U', value_U)
    assign_value(ws, new_row, 'V', value_V)
    assign_value(ws, new_row, 'W', value_W)

    assign_value(ws, new_row, 'X', value_X)
    assign_value(ws, new_row, 'Y', value_Y)
    assign_value(ws, new_row, 'Z', value_z)
    assign_value(ws, new_row, 'AA', value_AA)
    assign_value(ws, new_row, 'AB', value_AB)
    assign_value(ws, new_row, 'AD', value_AD)
    assign_value(ws, new_row, 'AE', value_AE)
    assign_value(ws, new_row, 'AF', value_AF)
    assign_value(ws, new_row, 'AG', value_AG)
    assign_value(ws, new_row, 'AH', value_AH)
    assign_value(ws, new_row, 'AI', value_AI)
    assign_value(ws, new_row, 'AJ', value_AJ)
    assign_value(ws, new_row, 'AK', value_AK)
    assign_value(ws, new_row, 'AL', value_AL)
    assign_value(ws, new_row, 'AM', value_AM)
    # assign_value(ws, row_AN, 'AN',value_AN)

def save_xi_worksheet(ws):
    rows = int(ws.max_row)
    ###############
    # 查数据
    
    data17 = query_xi_back_gun()
    data18 = query_xi_ecpm()
    data19 = query_xi_ecpm(country_us)


    # 数据计算
    value_G_1 = round_sig(data18['昨天'])    # G列 昨日ecpm
    value_G_2 = round_sig(data18['前天'])    # G列 前日ecpm
    value_H_1 = round_sig(data19['昨天'])    # F列 昨日US-ecpm
    value_H_2 = round_sig(data19['前天'])    # F列 前日US-ecpm
    value_AN = round_sig(data17['num_video_played']/data17['people_num_watch_video'])      # AN列M美国fb用户24小时视频数

    
    
    # 切换列数
    # new_row = rows + 1
    row_G = rows-1
    row_F = rows-1    
    # row_I = new_row-1
    # row_J = new_row-3
    # row_K = new_row-7
    row_AN= rows-2

    # 填数据到excel

    assign_value(ws, rows, 'G', value_G_1)
    assign_value(ws, row_G, 'G', value_G_2)
    assign_value(ws, rows, 'H', value_H_1)
    assign_value(ws, row_F, 'H', value_H_2)
    assign_value(ws, row_AN, 'AN',value_AN)

def round_sig(x, sig=2):
    return round(x, sig-int(floor(log10(abs(x)))))


def save_to_ios(wb):
    stat_platform_ios()
    ws = wb["iOS"]   
    save_umeng_worksheet(ws)
    save_xi_worksheet(ws)
    # save_to_worksheet(ws)





def save_to_android(wb):
    stat_platform_android()
    ws = wb["安卓"]
    save_umeng_worksheet(ws)
    save_xi_worksheet(ws)
    # save_to_worksheet(ws)



def open_excel(filename):
    wb = openpyxl.load_workbook(filename)
    return wb


def save_excel(wb):
    # Save the file
    wb.save(save_file_name)

def test():
    filename = "TBC3每日数据20190912.xlsx"
    wb = open_excel(filename)
    ws = wb["iOS"]
    rows = int(ws.max_row)
    new_row = rows + 1
    assign_value(ws, new_row, 'A', yesterday)
    assign_value(ws, new_row, 'B', 'iOS')
    assign_value(ws, new_row, 'C', 123456789)
    assign_value(ws, new_row, 'D', 789456123)
    assign_value(ws, new_row, 'I', 45)
    save_excel(wb)


# test()
