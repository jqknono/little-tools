import pandas
from datetime import date, datetime, timedelta
from xlutils.copy import copy
import openpyxl
from functions import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from copy import copy

# df = pandas.read_excel("工作簿1.xlsx",sheet_name="iOS")
# df.to_excel("工作簿2.xlsx",sheet_name="iOS",)

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

filename = "工作簿1.xlsx"
wb = openpyxl.load_workbook(filename)
# grab the active worksheet
ws = wb.active

columns = ws.max_column
rows = ws.max_row

def assign_value(row,column,value):
    ws[f'{column}{rows+1}'] = value
    ws[f'{column}{rows+1}'].font = copy(ws[f'{column}{rows-10}'].font)
    ws[f'{column}{rows+1}'].color = copy(ws[f'{column}{rows-10}'].color)
    ws[f'{column}{rows+1}'].PatternFill = copy(ws[f'{column}{rows-10}'].PatternFill)
    ws[f'{column}{rows+1}'].Border = copy(ws[f'{column}{rows-10}'].Border)
    

today = datetime.today()
yesterday = (today - timedelta(days=1)).strftime('%Y/%m/%d')
save_file_name = (today - timedelta(days=1)).strftime('%Y-%m-%d')





###############
# 查数据
data = query_active_users()
###############

###############
# 填数据到excel
#ws.append([   yesterday, 'iOS', data["昨天"][1], data['昨天'][0]         ])


#ws[f'I{rows-1}'] = "fanfan"

assign_value(rows, 'C',data['昨天'][1])
assign_value(rows, 'D',data['昨天'][0])
#assign_value(rows, 'E',data['昨天'][0])


###############

ws[f'A{rows+1}'].font = copy(ws[f'A{rows}'].font)
ws[f'B{rows+1}'].font = copy(ws[f'B{rows}'].font)
ws[f'C{rows+1}'].font = copy(ws[f'C{rows}'].font)

# Save the file
wb.save(f"TBC3每日数据{save_file_name}.xlsx")
